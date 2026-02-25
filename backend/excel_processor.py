"""
Excel processor for UAE bank statements without pandas dependency
Uses openpyxl directly for better Excel compatibility
"""
import json
import re
from datetime import datetime
import os
import openai

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Set OpenAI API key
openai.api_key = os.getenv('OPENAI_API_KEY')

class UAEBankExcelProcessor:
    def __init__(self):
        self.bank_patterns = {
            'ADCB': ['abu dhabi commercial bank', 'adcb'],
            'FAB': ['first abu dhabi bank', 'fab'],
            'ENBD': ['emirates nbd', 'enbd'],
            'MASHREQ': ['mashreq bank', 'mashreq'],
            'CBD': ['commercial bank of dubai', 'cbd'],
            'HSBC': ['hsbc'],
            'RAKBANK': ['rak bank', 'rakbank'],
            'ADIB': ['abu dhabi islamic bank', 'adib'],
            'BOA': ['bank of america', 'bofa', 'boa', 'b of a'],
            'CHASE': ['chase', 'jp morgan chase', 'jpmorgan'],
            'WELLS': ['wells fargo', 'wells'],
            'CITI': ['citibank', 'citi']
        }

        # UAE IBFT / mobile-banking transfer reference pattern (e.g. 530P79B7E39A6295)
        # Matches 8+ hex/alphanumeric strings that look like bank reference IDs
        self._ibft_ref_pattern = re.compile(
            r'^[0-9a-f]{6,}[0-9a-f\-]*$|^[0-9]{3}[a-z][0-9a-f]{8,}',
            re.IGNORECASE
        )

        self.categories = {
            'Internal Transfer': [
                'mobn transfer', 'mobn', 'internal transfer', 'own account transfer',
                'self transfer', 'ibft', 'neft', 'rtgs', 'imps', 'upi transfer',
                'between accounts', 'acc transfer', 'account transfer',
                'transfer from', 'transfer to', 'online transfer', 'mobile transfer',
                'internet banking transfer', 'e-transfer', 'wire to self',
                'salary transfer own', 'al ansari', 'uae exchange transfer to self',
            ],
            'Salary & Income': [
                'salary', 'payroll', 'wages', 'income', 'direct deposit', 'paycheck',
                'pay slip', 'bonus', 'commission', 'stipend', 'allowance',
                'reimbursement', 'refund', 'credit from', 'deposit from', 'payment from',
                'received from', 'dividend', 'interest credit', 'profit share',
            ],
            'Food & Beverage': [
                'carrefour', 'lulu', 'spinneys', 'choithrams', 'union coop', 'waitrose',
                'restaurant', 'cafe', 'kfc', 'mcdonald', 'pizza', 'subway', 'dominos',
                'starbucks', 'costa', 'dunkin', 'burger', 'food', 'dining', 'eat',
                'grocery', 'supermarket', 'hypermarket', 'bakery', 'deli', 'bistro',
                'catering', 'takeaway', 'delivery', 'zomato', 'talabat', 'deliveroo',
                'doordash', 'grubhub', 'uber eats', 'postmates', 'chipotle', 'panera',
                'whole foods', 'trader joe', 'safeway', 'kroger',
                'donuts', 'coffee', 'espresso', 'bacio di latte', 'butchery', 'zinque',
                'ralphs', 'albertsons', 'vons', 'pavilions', 'publix', 'wegmans', 'harris teeter',
                'noon food', 'kcal', 'just eat', 'hunger station', 'marsool', 'mrsool',
                'jahez', 'tikka', 'shawarma', 'manoushe', 'falafel', 'hummus',
                'baqala', 'chaat', 'karak', 'cafetera',
                'moongoat', 'dutch bros', "peet's coffee", 'peets coffee', 'blue bottle',
                'instacart', 'fresh direct', 'imperfect foods', 'misfit market',
                'chick-fil-a', 'chickfila', 'taco bell', 'five guys', 'shake shack',
                'in-n-out', 'whataburger', 'sonic drive', 'dairy queen', 'popeyes',
                'olive garden', 'red lobster', 'applebees', 'chilis', 'ihop', 'dennys',
                'wingstop', 'buffalo wild wings', 'raising canes', 'culvers',
            ],
            'Transportation & Logistics': [
                'adnoc', 'eppco', 'enoc', 'petrol', 'fuel', 'gas', 'gasoline',
                'taxi', 'uber', 'careem', 'metro', 'bus', 'rta', 'parking',
                'salik', 'toll', 'car wash', 'transport', 'emirates', 'etihad',
                'flydubai', 'air arabia', 'airline', 'flight', 'airport',
                'chevron', 'shell', 'bp', '76', 'exxon', 'mobil', 'lyft',
                'parking services', 'toll roads', 'valet', 'bolt', 'indrive', 'yango',
                'aramex', 'dhl', 'fedex', 'ups', 'courier', 'shipping', 'freight',
            ],
            'Retail & Shopping': [
                'mall', 'centrepoint', 'max', 'home centre', 'ikea', 'ace',
                'sharaf dg', 'jumbo', 'electronics', 'clothing', 'fashion',
                'shop', 'store', 'retail', 'amazon', 'noon', 'souq', 'namshi',
                'h&m', 'zara', 'nike', 'adidas', 'apple store', 'samsung', 'virgin',
                'uniqlo', 'target', 'walmart', 'costco', 'best buy', 'macy',
                'nordstrom', 'kohls', 'tj maxx', 'ross', 'marshalls',
                'shein', 'bosta',
            ],
            'Healthcare': [
                'hospital', 'clinic', 'pharmacy', 'medical', 'doctor', 'health',
                'dental', 'medicare', 'aster', 'nmc', 'mediclinic', 'life pharmacy',
                'boots', 'aster pharmacy', 'day today pharmacy',
                'seha', 'dha', 'nmc healthcare', 'thumbay', 'daman',
                'salon', 'spa', 'barbershop', 'beauty', 'cosmetics', 'skincare',
                'haircut', 'manicure', 'pedicure', 'massage', 'gym', 'fitness',
                '24hourfitness', 'planet fitness', 'la fitness', 'crunch', 'equinox',
                'flex fitness', 'orangetheory',
            ],
            'Utilities': [
                'dewa', 'addc', 'sewa', 'fewa', 'etisalat', 'du', 'internet',
                'mobile', 'telecom', 'electricity', 'water', 'utility', 'bill',
                'wifi', 'broadband', 'phone bill', 'electric bill',
            ],
            'Entertainment & Media': [
                'cinema', 'movie', 'vox', 'reel', 'osn', 'gaming',
                'entertainment', 'park', 'beach', 'attraction', 'ticket', 'event',
                'youtube', 'disney', 'hulu', 'shahid',
                'disneyland', 'balloon museum', 'sawdust festival', 'museum', 'amusement',
            ],
            'Technology': [
                'netflix', 'spotify', 'youtube premium', 'amazon prime', 'disney+',
                'adobe', 'microsoft', 'google', 'icloud', 'dropbox', 'zoom',
                'subscription', 'monthly', 'annual', 'recurring', 'saas',
                'software', 'app store', 'play store', 'itunes', 'office 365',
                'hbo', 'peacock', 'paramount+', 'apple tv', 'apple one',
                'slack', 'github', 'figma', 'notion', 'canva', 'grammarly',
                'dashlane', '1password', 'nordvpn', 'expressvpn', 'duolingo',
                'headspace', 'calm', 'audible', 'kindle unlimited',
            ],
            'Finance & Banking': [
                'atm', 'cash withdrawal', 'withdrawal', 'atm withdrawal',
                'cash advance', 'atm fee', 'withdrawal fee',
                'bank fee', 'service charge', 'maintenance fee', 'overdraft', 'annual fee',
                'interest charge', 'finance charge', 'late payment fee',
                'wire transfer', 'remittance', 'foreign exchange', 'forex', 'currency exchange',
                'uae exchange', 'al ansari exchange', 'lulu exchange', 'wall street exchange',
                'zelle', 'venmo', 'paypal', 'western union', 'moneygram',
                'loan repayment', 'emi payment', 'credit card payment', 'insurance premium',
                'ach debit', 'ach credit', 'direct debit', 'direct deposit',
                'pos purchase', 'checkcard', 'wire to', 'wire from',
                'cash app', 'apple pay', 'google pay', 'samsung pay',
                'tabby', 'spotii', 'cashew', 'tamara',
            ]
        }

    def detect_bank(self, text):
        """Detect bank from text content"""
        text_lower = text.lower()
        for bank_code, patterns in self.bank_patterns.items():
            for pattern in patterns:
                if pattern in text_lower:
                    return {
                        'ADCB': 'Abu Dhabi Commercial Bank',
                        'FAB': 'First Abu Dhabi Bank',
                        'ENBD': 'Emirates NBD',
                        'MASHREQ': 'Mashreq Bank',
                        'CBD': 'Commercial Bank of Dubai',
                        'HSBC': 'HSBC Bank Middle East',
                        'RAKBANK': 'RAKBank',
                        'ADIB': 'Abu Dhabi Islamic Bank',
                        'BOA': 'Bank of America',
                        'CHASE': 'Chase Bank',
                        'WELLS': 'Wells Fargo',
                        'CITI': 'Citibank'
                    }.get(bank_code, f'{bank_code} Bank')
        return 'Unknown Bank'

    def _is_ibft_reference(self, description):
        """Return True if description looks like a UAE bank IBFT reference number
           e.g. '530P79B7E39A6295 - M', 'FT25239XYZABC123'
        """
        # Strip trailing reference suffix like '- M', '- REF', etc.
        clean = re.sub(r'\s*[-–]\s*\w{0,5}\s*$', '', description.strip())
        # Match if the cleaned string is mostly hex/alphanumeric (reference ID)
        return bool(re.match(r'^[0-9A-F]{8,}$', clean, re.IGNORECASE))

    def extract_merchant_name(self, description):
        """Extract a clean merchant name from a raw bank transaction description.

        Strips US bank prefixes (CHECKCARD MMDD, MOBILE PURCHASE MMDD, etc.),
        trailing location info, long reference numbers, and RECURRING suffixes.
        Returns a human-readable merchant name, title-cased, max 60 chars.
        """
        s = description.strip()

        # Step 1: Strip known US bank transaction prefixes (with optional 4-digit date MMDD)
        prefix_patterns = [
            r'^CHECKCARD\s+\d{4}\s+',
            r'^MOBILE\s+PURCHASE\s+\d{4}\s+',
            r'^POS\s+PURCHASE\s+\d{4}\s+',
            r'^RECURRING\s+PURCHASE\s+\d{4}\s+',
            r'^ONLINE\s+PURCHASE\s+\d{4}\s+',
            r'^DEBIT\s+CARD\s+PURCHASE\s+\d{4}\s+',
            r'^DEBIT\s+PURCHASE\s+\d{4}\s+',
            r'^ACH\s+DEBIT\s+',
            r'^ACH\s+CREDIT\s+',
            r'^WIRE\s+TRANSFER\s+(TO|FROM)\s+',
            r'^WIRE\s+(TO|FROM)\s+',
            r'^ZELLE\s+(TO|FROM|PAYMENT)\s+',
            r'^VENMO\s+PAYMENT\s+',
            r'^PAYPAL\s+(TRANSFER\s+)?',
            r'^DIRECT\s+(DEBIT|DEPOSIT)\s+',
            r'^ATM\s+WITHDRAWAL\s+',
            r'^CHECKCARD\s+',
            r'^MOBILE\s+PURCHASE\s+',
            r'^POS\s+PURCHASE\s+',
        ]
        for pattern in prefix_patterns:
            s = re.sub(pattern, '', s, flags=re.IGNORECASE).strip()

        # Step 2: Handle SQ * (Square payments) — extract merchant name after SQ *
        sq_match = re.match(r'^SQ\s*\*\s*(.+)', s, re.IGNORECASE)
        if sq_match:
            s = sq_match.group(1).strip()

        # Step 3: Remove trailing US location info (City ST 12345 or just ST 12345)
        s = re.sub(r'\s+[A-Z]{2}\s+\d{5}(-\d{4})?$', '', s).strip()
        s = re.sub(r'\s+[A-Z]{2}$', '', s).strip()

        # Step 4: Remove long numeric reference strings (8+ digits)
        s = re.sub(r'\s+\d{8,}', '', s).strip()
        # Remove long hex/alphanumeric reference codes (10+ chars)
        s = re.sub(r'\b[0-9A-Fa-f]{10,}\b', '', s).strip()

        # Step 5: Remove RECURRING / RECURRING CHARGE suffix
        s = re.sub(r'\s+RECURRING(\s+CHARGE)?$', '', s, flags=re.IGNORECASE).strip()

        # Step 6: Remove store numbers (#0299, #1234)
        s = re.sub(r'\s+#\d+.*$', '', s).strip()

        # Step 7: Remove trailing "N 8445052993 CA..." style junk after first word cluster
        # Pattern: word(s) then space then number then space then more junk
        s = re.sub(r'^([A-Za-z][A-Za-z0-9\s\.\&\'\-]{1,40}?)\s+\d+\s+[A-Z].*$', r'\1', s).strip()

        # Step 8: Title-case and trim to reasonable length
        result = s.title().strip()
        return result[:60] if result else description[:60]

    def categorize_transaction(self, description):
        """Enhanced categorization with priority matching and better logic.
        Pass the MerchantName (cleaned) for best accuracy, or raw Description as fallback.
        """
        desc_lower = description.lower().strip()

        # --- Handle SQ * (Square payments) — categorize by the merchant after SQ * ---
        sq_match = re.match(r'^sq\s*\*\s*(.+)', desc_lower)
        if sq_match:
            merchant_after_sq = sq_match.group(1).strip()
            cat = self.categorize_transaction(merchant_after_sq)
            if cat != 'Other':
                return cat
            return 'Retail & Shopping'  # Default for unknown Square merchants

        # --- UAE-specific pre-checks ---
        # MOBN = Mobile Banking transfer; hex refs = IBFT/bank transfer
        if ('mobn' in desc_lower or
                desc_lower.startswith('mobn transfer') or
                self._is_ibft_reference(description)):
            return 'Internal Transfer'

        # Salary / incoming credit signals
        if any(kw in desc_lower for kw in ['salary', 'payroll', 'paycheck', 'direct deposit',
                                             'wages', 'commission payment', 'bonus credit']):
            return 'Salary & Income'

        # Priority-based categorization for better accuracy
        category_priority = [
            'Internal Transfer',
            'Salary & Income',
            'Finance & Banking',
            'Technology',
            'Food & Beverage',
            'Transportation & Logistics',
            'Healthcare',
            'Utilities',
            'Entertainment & Media',
            'Retail & Shopping',
        ]

        # Check for exact keyword matches in priority order
        for category in category_priority:
            keywords = self.categories.get(category, [])
            for keyword in keywords:
                if keyword in desc_lower:
                    if category == 'Finance & Banking':
                        if any(term in desc_lower for term in ['atm', 'withdrawal', 'cash', 'fee',
                                'charge', 'interest', 'maintenance', 'remittance', 'exchange',
                                'zelle', 'venmo', 'paypal', 'western union', 'forex', 'overdraft',
                                'tabby', 'spotii', 'tamara']):
                            return category
                    elif category == 'Technology':
                        if any(term in desc_lower for term in ['subscription', 'monthly', 'netflix',
                                'spotify', 'prime', 'office', 'adobe', 'google', 'microsoft', 'icloud']):
                            return category
                    else:
                        return category

        # Additional pattern matching for common transaction types
        if any(word in desc_lower for word in ['taxi', 'uber', 'careem', 'rta', 'bolt', 'indrive', 'yango']):
            return 'Transportation & Logistics'

        if any(word in desc_lower for word in ['restaurant', 'cafe', 'food', 'dining', 'eat',
                                                'talabat', 'zomato', 'deliveroo', 'mrsool', 'marsool']):
            return 'Food & Beverage'

        if any(word in desc_lower for word in ['mall', 'shop', 'store', 'retail', 'noon', 'amazon']):
            return 'Retail & Shopping'

        # Merchant shorthand map
        merchant_patterns = {
            'careem': 'Transportation & Logistics',
            'talabat': 'Food & Beverage',
            'zomato': 'Food & Beverage',
            'mrsool': 'Food & Beverage',
            'marsool': 'Food & Beverage',
            'netflix': 'Technology',
            'spotify': 'Technology',
            'amazon': 'Retail & Shopping',
            'noon': 'Retail & Shopping',
            'namshi': 'Retail & Shopping',
            'adnoc': 'Transportation & Logistics',
            'enoc': 'Transportation & Logistics',
            'eppco': 'Transportation & Logistics',
            'salik': 'Transportation & Logistics',
            'rta': 'Transportation & Logistics',
            'dewa': 'Utilities',
            'etisalat': 'Utilities',
            'e&': 'Utilities',
            'du telecom': 'Utilities',
            'addc': 'Utilities',
            'sewa': 'Utilities',
            'tabby': 'Finance & Banking',
            'spotii': 'Finance & Banking',
            'tamara': 'Finance & Banking',
        }

        for merchant, cat in merchant_patterns.items():
            if merchant in desc_lower:
                return cat

        return 'Other'

    def ai_categorize_transactions(self, transactions):
        """Use AI to intelligently categorize transactions in batches"""
        if not openai.api_key or not transactions:
            return transactions

        # Define available categories — exact frontend sector names (no mapping needed)
        categories = [
            'Internal Transfer',
            'Salary & Income',
            'Food & Beverage',
            'Transportation & Logistics',
            'Retail & Shopping',
            'Healthcare',
            'Utilities',
            'Entertainment & Media',
            'Technology',
            'Finance & Banking',
            'Travel & Tourism',
            'Professional Services',
            'Real Estate',
            'Education',
            'Other',
        ]

        try:
            # Process in batches of 50 transactions
            batch_size = 50
            categorized_transactions = []

            for i in range(0, len(transactions), batch_size):
                batch = transactions[i:i + batch_size]

                # Prepare batch for AI — use clean MerchantName when available
                descriptions = [
                    f"{idx}. {tx.get('MerchantName') or tx['Description']}"
                    for idx, tx in enumerate(batch, start=i)
                ]

                prompt = f"""You are a UAE/Gulf bank statement categorization expert. Categorize each transaction into exactly one category. READ CAREFULLY — many descriptions are UAE bank reference codes.

Available categories: {', '.join(categories)}

STRICT RULES:

1. Internal Transfer — BETWEEN OWN ACCOUNTS (most important rule):
   ✓ Any description that is a hex/alphanumeric reference ID (e.g. "530P79B7E39A6295 - M", "FT25239ABC123", "IBFT REF 12345")
   ✓ "MOBN TRANSFER FROM", "MOBN TRANSFER TO" (MOBN = Mobile Banking UAE)
   ✓ "Own account transfer", "Transfer between accounts", "Online transfer"
   ✓ Long alphanumeric codes that look like bank reference numbers
   ✗ NOT transfers to other people or businesses

2. Salary & Income — MONEY RECEIVED:
   ✓ "SALARY", "PAYROLL", "WAGES", "COMMISSION", "BONUS"
   ✓ Direct deposit credits, employer payments
   ✓ Refunds and reimbursements

3. Food & Beverage — RESTAURANTS, CAFES, GROCERIES:
   ✓ Carrefour, Lulu, Spinneys, Choithrams, Union Coop, Baqala, Chaat, Karak
   ✓ Talabat, Zomato, Deliveroo, Mrsool, Marsool, Hunger Station, Jahez
   ✓ McDonald's, KFC, Pizza Hut, Subway, Starbucks, Costa
   ✓ Any restaurant, cafe, bakery, food delivery

4. Transportation & Logistics — GAS, PARKING, RIDESHARE, SALIK:
   ✓ ADNOC, ENOC, EPPCO (fuel stations)
   ✓ Careem, Uber (rides), Bolt, inDrive, Yango, RTA, Salik (Dubai toll)
   ✓ Parking, airport transfers, Emirates, Flydubai, Air Arabia (short trips)
   ✓ DHL, FedEx, Aramex, UPS (courier/shipping)

5. Utilities — DEWA, ETISALAT, DU:
   ✓ DEWA, ADDC, SEWA, FEWA (electricity/water)
   ✓ Etisalat / e&, du Telecom (phone/internet)
   ✓ Any utility bill payment

6. Technology — SOFTWARE SUBSCRIPTIONS & DIGITAL:
   ✓ Netflix, Spotify, Disney+, OSN, Shahid
   ✓ Adobe, Microsoft 365, iCloud, Dropbox, Zoom
   ✓ Any monthly/annual software subscription

7. Retail & Shopping:
   ✓ Noon, Amazon, Namshi, Souq, Shein
   ✓ H&M, Zara, Nike, Adidas, Sharaf DG, Jumbo
   ✓ Mall purchases, IKEA, Home Centre, Target, Walmart

8. Healthcare:
   ✓ Aster, NMC, Mediclinic, Life Pharmacy, Boots, SEHA, DHA
   ✓ Doctors, dentists, hospitals, clinics, pharmacies
   ✓ Gyms, salons, spas, fitness centers

9. Finance & Banking — BANK FEES & ATM:
   ✓ ATM Withdrawal, Cash Advance
   ✓ Annual fee, service charge, overdraft fee, interest charge
   ✓ UAE Exchange, Al Ansari (money exchange/remittance)
   ✓ Zelle, Western Union, MoneyGram, Tabby, Spotii, Tamara
   ✗ NOT own-account transfers (those are Internal Transfer)

10. Entertainment & Media:
   ✓ Vox Cinemas, IMAX, concerts, theme parks, VOX, Expo
   ✓ Gaming, sports events, museums, attractions

11. Travel & Tourism:
   ✓ Hotels, Airbnb, Booking.com, Agoda
   ✓ Long-haul airlines, rental cars, travel agencies

12. Professional Services:
   ✓ Consulting, legal, accounting, HR, advisory, audit firms

13. Education:
   ✓ Schools, universities, tuition, courses, training

14. Real Estate:
   ✓ Rent payment, property management, real estate agencies

15. Other — ONLY if truly cannot be identified

CRITICAL UAE EXAMPLES:
- "530P79B7E39A6295 - M" → Internal Transfer (hex bank reference ID)
- "MOBN TRANSFER FROM 12345" → Internal Transfer (Mobile Banking)
- "FT25239ABC123456" → Internal Transfer (Faster Payment reference)
- "DEWA BILL PAYMENT" → Utilities
- "SALIK RECHARGE" → Transportation & Logistics
- "TALABAT AE" → Food & Beverage
- "CARREFOUR HYPERMARKET" → Food & Beverage
- "SALARY - ACME LLC" → Salary & Income
- "ADNOC DISTRIBUTION" → Transportation & Logistics
- "ETISALAT POSTPAID" → Utilities
- "VOX CINEMAS" → Entertainment & Media
- "AL ANSARI EXCHANGE" → Finance & Banking
- "ATM WITHDRAWAL" → Finance & Banking
- "TARGET ST 4255" → Retail & Shopping
- "RALPHS #0299" → Food & Beverage (grocery store)
- "Netflix" → Technology
- "Verizon Wireless" → Utilities (phone service)
- "AMAZON.COM" → Retail & Shopping
- "TABBY" → Finance & Banking (BNPL)

Transactions to categorize:
{chr(10).join(descriptions)}

Respond ONLY with a JSON array of category names in the EXACT order of transactions, nothing else.
Example format: ["Food & Beverage", "Transportation & Logistics", "Retail & Shopping"]"""

                try:
                    response = openai.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {
                                "role": "system",
                                "content": "You are a financial categorization expert. Respond only with a JSON array of categories, no additional text."
                            },
                            {"role": "user", "content": prompt}
                        ],
                        max_tokens=800,
                        temperature=0.1
                    )

                    ai_response = response.choices[0].message.content.strip()

                    # Remove markdown if present
                    if ai_response.startswith('```json'):
                        ai_response = ai_response.replace('```json', '').replace('```', '').strip()
                    elif ai_response.startswith('```'):
                        ai_response = ai_response.replace('```', '').strip()

                    ai_categories = json.loads(ai_response)

                    # Pad with rule-based fallbacks if AI returned fewer than batch size
                    while len(ai_categories) < len(batch):
                        ai_categories.append(batch[len(ai_categories)].get('Category', 'Other'))

                    # Apply AI categories to batch
                    for idx, tx in enumerate(batch):
                        if idx < len(ai_categories):
                            tx['Category'] = ai_categories[idx]
                            # Subcategory = full category name (not just first word)
                            tx['Subcategory'] = ai_categories[idx]
                        categorized_transactions.append(tx)

                except Exception as e:
                    print(f"AI categorization error for batch: {str(e)}")
                    # Fallback to rule-based for this batch
                    for tx in batch:
                        categorized_transactions.append(tx)

            return categorized_transactions

        except Exception as e:
            print(f"AI categorization failed: {str(e)}")
            return transactions

    def extract_bank_info(self, worksheet):
        """Extract bank information from worksheet"""
        bank_info = {
            'bank_name': 'UAE Bank',
            'account_holder': 'Account Holder',
            'account_number': 'XXXX-XXXX-XXXX',
            'currency': 'AED'
        }

        # Check first 20 rows for bank information
        for row in range(1, min(21, worksheet.max_row + 1)):
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell_value = str(worksheet.cell(row, col).value or '').upper()

                # Detect bank
                bank_name = self.detect_bank(cell_value)
                if bank_name not in ['UAE Bank', 'Unknown Bank']:
                    bank_info['bank_name'] = bank_name

                # Look for account holder (names are usually 2-4 words, mixed case)
                if len(cell_value.split()) >= 2 and len(cell_value.split()) <= 4:
                    if not any(char.isdigit() for char in cell_value):
                        if any(word in cell_value.lower() for word in ['mr', 'ms', 'mrs', 'dr']):
                            bank_info['account_holder'] = cell_value.title()

                # Look for account numbers
                if re.match(r'^\d{10,16}$', cell_value.replace(' ', '').replace('-', '')):
                    masked = cell_value[:4] + '-' + '*' * 4 + '-' + cell_value[-4:]
                    bank_info['account_number'] = masked

        return bank_info

    def find_data_headers(self, worksheet):
        """Find the row with data headers and detect column mappings"""
        header_keywords = ['date', 'amount', 'description', 'particular', 'narration', 'debit', 'credit', 'balance', 'type', 'reference']

        for row in range(1, min(20, worksheet.max_row + 1)):
            row_data = {}
            header_count = 0

            for col in range(1, min(worksheet.max_column + 1, 10)):  # Check up to 10 columns
                cell_value = str(worksheet.cell(row, col).value or '').lower().strip()

                # Map common header variations
                if cell_value and any(keyword in cell_value for keyword in header_keywords):
                    header_count += 1

                    # Detect column types
                    if 'date' in cell_value:
                        row_data['date_col'] = col
                    elif 'description' in cell_value or 'particular' in cell_value or 'narration' in cell_value:
                        row_data['description_col'] = col
                    elif 'type' in cell_value:
                        row_data['type_col'] = col
                    elif 'reference' in cell_value or 'ref' in cell_value:
                        row_data['reference_col'] = col
                    elif 'debit' in cell_value:
                        row_data['debit_col'] = col
                    elif 'credit' in cell_value:
                        row_data['credit_col'] = col
                    elif 'amount' in cell_value:
                        row_data['amount_col'] = col

            # If we found at least 3 header columns, this is likely the header row
            if header_count >= 3:
                row_data['header_row'] = row
                return row_data

        # Default mapping if no headers found
        return {'header_row': 1, 'date_col': 1, 'description_col': 2, 'debit_col': 4, 'credit_col': 5}

    def detect_date_format(self, date_string, currency='AED'):
        """Detect and parse date from various formats.

        Uses the currency hint to resolve ambiguous D/M/YYYY vs M/D/YYYY dates:
          - AED (UAE) banks use DD/MM/YYYY
          - USD (US) banks use MM/DD/YYYY
        If only one interpretation is valid (one component > 12), that wins regardless.
        """
        if isinstance(date_string, datetime):
            return date_string.strftime('%Y-%m-%d')

        date_str = str(date_string).strip()

        # MM/DD/YY — US 2-digit year (e.g. 12/06/24 → 2024-12-06)
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', date_str)
        if m:
            month, day, year = m.group(1), m.group(2), m.group(3)
            full_year = '20' + year if int(year) < 50 else '19' + year
            return f"{full_year}-{month.zfill(2)}-{day.zfill(2)}"

        # YYYY-MM-DD — ISO format
        m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', date_str)
        if m:
            return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"

        # D/M/YYYY or M/D/YYYY with 4-digit year — resolve using heuristics then currency
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', date_str)
        if m:
            first, second, year = int(m.group(1)), int(m.group(2)), m.group(3)
            # Unambiguous: first component > 12 → must be the day (DD/MM/YYYY)
            if first > 12:
                return f"{year}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
            # Unambiguous: second component > 12 → must be the day (MM/DD/YYYY)
            if second > 12:
                return f"{year}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
            # Both ≤ 12: ambiguous — use currency hint
            if currency == 'USD':
                # US banks: MM/DD/YYYY
                return f"{year}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
            else:
                # UAE/other banks: DD/MM/YYYY
                return f"{year}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"

        # Fallback to current date if parsing fails
        return datetime.now().strftime('%Y-%m-%d')

    def process_excel_file(self, file_content):
        """Dynamic Excel processor - automatically detects and adapts to different Excel formats"""
        if not EXCEL_AVAILABLE:
            return None, "openpyxl not available - using demo data"

        try:
            # Load workbook
            workbook = load_workbook(file_content)
            print(f"[*] Processing Excel file with {len(workbook.sheetnames)} sheets")

            # Extract bank info from Account Info sheet
            bank_info = {'bank_name': 'Unknown Bank', 'account_holder': 'Account Holder', 'account_number': 'XXXX-XXXX-XXXX', 'currency': 'USD'}
            if 'Account Info' in workbook.sheetnames:
                info_sheet = workbook['Account Info']
                for row in range(1, info_sheet.max_row + 1):
                    field = str(info_sheet.cell(row, 1).value or '').lower()
                    value = str(info_sheet.cell(row, 2).value or '')
                    if 'account holder' in field and value:
                        bank_info['account_holder'] = value
                    elif 'account number' in field and value:
                        bank_info['account_number'] = value
                    elif 'bank name' in field and value:
                        detected_bank = self.detect_bank(value)
                        if detected_bank != 'Unknown Bank':
                            bank_info['bank_name'] = detected_bank
                            # Set currency based on bank
                            if 'America' in detected_bank or 'Chase' in detected_bank or 'Wells' in detected_bank or 'Citi' in detected_bank:
                                bank_info['currency'] = 'USD'
                            else:
                                bank_info['currency'] = 'AED'

            print(f"[BANK] {bank_info['bank_name']} | Currency: {bank_info['currency']}")

            # Process transaction sheets (all sheets except Account Info)
            all_transactions = []
            transaction_sheets = [sheet for sheet in workbook.sheetnames if sheet != 'Account Info']

            for sheet_name in transaction_sheets:
                worksheet = workbook[sheet_name]
                print(f"\n[SHEET] Processing: {sheet_name}")

                # DYNAMIC HEADER DETECTION - automatically find headers and column mappings
                column_map = self.find_data_headers(worksheet)
                header_row = column_map.get('header_row', 1)
                date_col = column_map.get('date_col', 1)
                desc_col = column_map.get('description_col', 2)
                type_col = column_map.get('type_col', None)
                debit_col = column_map.get('debit_col', 4)
                credit_col = column_map.get('credit_col', 5)
                amount_col = column_map.get('amount_col', None)

                print(f"  [OK] Headers found at row {header_row}")
                print(f"  [OK] Columns: Date={date_col}, Desc={desc_col}, Type={type_col}, Debit={debit_col}, Credit={credit_col}")

                # Process transactions starting from row after headers
                sheet_transactions = 0
                for row in range(header_row + 1, worksheet.max_row + 1):
                    try:
                        # DYNAMIC CELL READING - read based on detected column positions
                        date_cell = worksheet.cell(row, date_col).value if date_col else None
                        description_cell = worksheet.cell(row, desc_col).value if desc_col else None
                        type_cell = worksheet.cell(row, type_col).value if type_col else None
                        debit_cell = worksheet.cell(row, debit_col).value if debit_col else None
                        credit_cell = worksheet.cell(row, credit_col).value if credit_col else None

                        # Skip empty rows or summary rows
                        if not date_cell:
                            continue
                        if any(keyword in str(date_cell).upper() for keyword in ['TOTAL', 'SUMMARY', 'BALANCE', 'OPENING', 'CLOSING', 'NET', 'FLOW']):
                            continue
                        if description_cell and any(keyword in str(description_cell).upper() for keyword in ['TOTAL', 'SUMMARY', 'MONTHLY', 'NET FLOW', 'NET DEBIT', 'NET CREDIT']):
                            continue

                        # INTELLIGENT DATE PARSING - automatically detect and parse date format
                        date_str = self.detect_date_format(date_cell, bank_info.get('currency', 'AED'))

                        # DYNAMIC AMOUNT EXTRACTION - handle both debit/credit and single amount columns
                        amount = 0
                        try:
                            if amount_col:
                                # Single amount column — use Type column for sign
                                raw_amount = float(str(worksheet.cell(row, amount_col).value or '0').replace(',', ''))
                                amount = raw_amount
                            else:
                                # Debit/Credit columns
                                if debit_cell and str(debit_cell).strip() and str(debit_cell).strip() not in ['', '-', 'None']:
                                    amount = -abs(float(str(debit_cell).replace(',', '')))  # Debits are negative
                                elif credit_cell and str(credit_cell).strip() and str(credit_cell).strip() not in ['', '-', 'None']:
                                    amount = abs(float(str(credit_cell).replace(',', '')))  # Credits are positive
                        except (ValueError, TypeError):
                            continue

                        # FIX 1A: Apply Type column to correct amount sign
                        # Type = "Withdrawal" / "Debit" / "DR" → negative
                        # Type = "Deposit" / "Credit" / "CR" → positive
                        if type_cell:
                            type_val = str(type_cell).strip().lower()
                            if type_val in ('withdrawal', 'debit', 'dr', 'debit card purchase'):
                                amount = -abs(amount)
                            elif type_val in ('deposit', 'credit', 'cr'):
                                amount = abs(amount)

                        # Skip if no amount
                        if amount == 0:
                            continue

                        # FIX 1B: Extract clean merchant name from raw description
                        raw_description = str(description_cell).strip() if description_cell else f"Transaction {len(all_transactions) + 1}"
                        merchant_name = self.extract_merchant_name(raw_description)

                        # Categorize using clean merchant name for better accuracy
                        category = self.categorize_transaction(merchant_name)
                        if category == 'Other':
                            # Fallback: try raw description too
                            category = self.categorize_transaction(raw_description)

                        # Preserve original type string for frontend
                        transaction_type = str(type_cell).strip() if type_cell else (
                            'Withdrawal' if amount < 0 else 'Deposit'
                        )

                        transaction = {
                            'Date': date_str,
                            'Amount': amount,
                            'Description': raw_description,
                            'MerchantName': merchant_name,
                            'Type': transaction_type,
                            'Category': category,
                            'Subcategory': category,
                        }
                        all_transactions.append(transaction)
                        sheet_transactions += 1

                    except Exception as e:
                        continue  # Skip problematic rows

                print(f"  [OK] Extracted {sheet_transactions} transactions from {sheet_name}")

            print(f"\n[SUCCESS] Total transactions extracted: {len(all_transactions)}")

            # Apply AI-powered categorization to improve accuracy
            if len(all_transactions) > 0:
                print(f"[AI] Applying AI categorization to {len(all_transactions)} transactions...")
                all_transactions = self.ai_categorize_transactions(all_transactions)
                print(f"[AI] AI categorization complete!")

            return {
                'transactions': all_transactions,
                'bank_info': bank_info,
                'total_rows': len(all_transactions),
                'processing_mode': 'Dynamic Excel Processing with AI Categorization'
            }, None

        except Exception as e:
            print(f"[ERROR] Error processing Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            return None, f"Error processing Excel file: {str(e)}"

# Global processor instance
processor = UAEBankExcelProcessor()